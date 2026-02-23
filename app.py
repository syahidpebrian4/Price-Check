import streamlit as st
import pytesseract
import cv2
import numpy as np
import pandas as pd
import re
import os
from PIL import Image, ImageDraw
from openpyxl import load_workbook
import io
import zipfile
from fuzzywuzzy import fuzz
import gc
import base64

# ================= CONFIG & DATABASE =================
FILE_PATH = "database/master_harga.xlsx"
SHEETS_TARGET = ["DF", "HBHC"]
SHEET_MASTER_IG = "IG"
COL_IG_NAME = "PRODNAME_IG"

st.set_page_config(page_title="Price Check", layout="wide", initial_sidebar_state="expanded")

# --- FUNGSI HELPER: LOGO BASE64 ---
def get_base64_image(image_path):
    if os.path.exists(image_path):
        with open(image_path, "rb") as img_file:
            return base64.b64encode(img_file.read()).decode()
    return None

# --- CSS CUSTOM ---
logo_b64 = get_base64_image("lotte_logo.png")
st.markdown(f"""
    <style>
        .custom-header {{
            position: fixed;
            top: 0; left: 0; width: 100%; height: 90px;
            background-color: white;
            display: flex; align-items: center;
            padding: 0 30px; border-bottom: 3px solid #eeeeee;
            z-index: 999999;
        }}
        .header-logo {{ height: 55px; margin-right: 25px; }}
        .header-title {{
            font-size: 42px; font-weight: 900;
            font-family: 'Arial Black', sans-serif; color: black; margin: 0;
        }}
        [data-testid="stSidebar"] {{
            background-color: #FF0000 !important;
            margin-top: 90px !important;
        }}
        .main .block-container {{ padding-top: 130px !important; }}
    </style>
    <div class="custom-header">
        <img src="data:image/png;base64,{logo_b64 if logo_b64 else ''}" class="header-logo">
        <h1 class="header-title">PRICE CHECK</h1>
    </div>
""", unsafe_allow_html=True)

# --- FUNGSI LOGIKA OCR ---
def clean_price_val(raw_str):
    if not raw_str: return 0
    table = str.maketrans('OISBEGZA', '01588624')
    text = str(raw_str).upper().translate(table)
    clean = re.sub(r'[^\d]', '', text)
    return int(clean) if clean else 0

def process_ocr_final(pil_image, master_product_names=None):
    img_np = cv2.cvtColor(np.array(pil_image), cv2.COLOR_RGB2BGR)
    scale = 2.0
    img_resized = cv2.resize(img_np, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    gray = cv2.cvtColor(img_resized, cv2.COLOR_BGR2GRAY)
    
    d = pytesseract.image_to_data(gray, output_type=pytesseract.Output.DICT, config=r'--oem 3 --psm 6')
    df_ocr = pd.DataFrame(d)
    df_ocr = df_ocr[df_ocr['text'].str.strip() != ""]
    df_ocr['text'] = df_ocr['text'].str.upper()

    # Logika pengelompokan baris teks
    df_ocr = df_ocr.sort_values(by=['top', 'left'])
    lines_data = []
    if not df_ocr.empty:
        current_top = df_ocr.iloc[0]['top']
        temp_words = []
        for _, row in df_ocr.iterrows():
            if row['top'] - current_top > 15:
                temp_words.sort(key=lambda x: x['left'])
                lines_data.append({
                    "text": " ".join([w['text'] for w in temp_words]),
                    "top": current_top,
                    "h": max([w['height'] for w in temp_words])
                })
                temp_words = [{'text': row['text'], 'left': row['left'], 'height': row['height']}]
                current_top = row['top']
            else:
                temp_words.append({'text': row['text'], 'left': row['left'], 'height': row['height']})
        lines_data.append({"text": " ".join([w['text'] for w in temp_words]), "top": current_top, "h": 10})

    lines_txt = [l['text'] for l in lines_data]
    full_text_single = " # ".join(lines_txt)

    prod_name, promo_desc = "N/A", "-"
    res = {"PCS": {"n": 0, "p": 0}, "CTN": {"n": 0, "p": 0}}
    
    # Matching Nama Produk
    if master_product_names:
        best_match, highest_score = "N/A", 0
        for ref_name in master_product_names:
            m_name = str(ref_name).upper()
            score = fuzz.partial_ratio(m_name, full_text_single)
            if score > 80 and score > highest_score:
                highest_score, best_match = score, m_name
        prod_name = best_match

    def extract_prices_from_line(line_text):
        line_text = re.sub(r'\(.*?\)|ISI\s*\d+', '', line_text)
        found_segments = re.split(r'RP|R9|BP|RD|P|R\s', line_text)
        found_prices = []
        for segment in found_segments:
            nums = re.findall(r'\d[\d\.,]+', segment)
            if nums:
                val = clean_price_val(nums[0])
                if 500 < val < 2000000:
                    found_prices.append(val)
        if not found_prices: return {"n": 0, "p": 0}
        n = found_prices[0]
        p = found_prices[1] if len(found_prices) >= 2 else found_prices[0]
        return {"n": n, "p": p}

    # Cari harga PCS & CTN
    for line in lines_txt:
        if any(k in line for k in ["PCS", "RCG", "BOX", "PCK", "PCH", "BTL"]) and "RP" in line:
            res_pcs = extract_prices_from_line(line)
            res["PCS"]["n"], res["PCS"]["p"] = res_pcs["n"], res_pcs["p"]
            break
    for line in lines_txt:
        if any(k in line for k in ["CTN", "KARTON", "DUS"]) and "RP" in line:
            res_ctn = extract_prices_from_line(line)
            res["CTN"]["n"], res["CTN"]["p"] = res_ctn["n"], res_ctn["p"]
            break

    return res["PCS"], res["CTN"], prod_name, "\n".join(lines_txt), pil_image, promo_desc

# ================= UI STREAMLIT =================
def norm(val):
    return str(val).replace(".0", "").replace(" ", "").strip().upper()

with st.sidebar:
    st.write("---")
    m_code = st.text_input("MASTER CODE").upper()
    date_inp = st.text_input("DATE (YYYYMMDD)").upper()
    week_inp = st.text_input("WEEK (1-52)")
    st.write("---")

files = st.file_uploader("UPLOAD GAMBAR", type=["jpg", "png", "jpeg"], accept_multiple_files=True)

if files and m_code and date_inp and week_inp:
    if os.path.exists(FILE_PATH):
        db_ig = pd.read_excel(FILE_PATH, sheet_name=SHEET_MASTER_IG)
        db_targets = {}
        for s in SHEETS_TARGET:
            df_tmp = pd.read_excel(FILE_PATH, sheet_name=s, header=3)
            df_tmp.columns = [str(c).strip().upper() for c in df_tmp.columns]
            db_targets[s] = df_tmp

        list_nama_master = db_ig[COL_IG_NAME].dropna().unique().tolist()
        final_list, zip_buffer = [], io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED) as zf:
            for f in files:
                with st.container(border=True):
                    img_pil = Image.open(f)
                    pcs, ctn, name, raw_txt, red_img, p_desc = process_ocr_final(img_pil, list_nama_master)
                    
                    # Cari ProdCode berdasarkan Nama yang didapat OCR
                    match_code, best_score = None, 0
                    for _, row in db_ig.iterrows():
                        db_name = str(row[COL_IG_NAME]).upper()
                        score = fuzz.partial_ratio(db_name, name)
                        if score > 75 and score > best_score:
                            best_score, match_code = score, norm(row["PRODCODE"])
                    
                    st.write(f"**File:** {f.name} | **Match:** {name} ({match_code})")
                    
                    if match_code:
                        for s_name, df_t in db_targets.items():
                            if "PRODCODE" in df_t.columns and "MASTER CODE" in df_t.columns:
                                match_row = df_t[(df_t["PRODCODE"].astype(str).apply(norm) == match_code) & 
                                                 (df_t["MASTER CODE"].astype(str).apply(norm) == norm(m_code))]
                                if not match_row.empty:
                                    final_list.append({
                                        "prodcode": match_code, "sheet": s_name, "index": match_row.index[0],
                                        "n_pcs": pcs['n'], "p_pcs": pcs['p'], "n_ctn": ctn['n'], "p_ctn": ctn['p'], "p_desc": p_desc
                                    })
                                    # Simpan foto ke ZIP
                                    img_byte = io.BytesIO()
                                    red_img.convert("RGB").save(img_byte, format="JPEG")
                                    zf.writestr(f"{match_code}.jpg", img_byte.getvalue())
                                    break
        
        if final_list:
            if st.button("💾 UPDATE DATABASE EXCEL"):
                wb = load_workbook(FILE_PATH)
                for r in final_list:
                    ws = wb[r['sheet']]
                    headers = [str(cell.value).strip().upper() for cell in ws[3]]
                    row_num = r['index'] + 4
                    
                    map_cols = {
                        "NORMAL COMPETITOR PRICE (PCS)": r['n_pcs'],
                        "PROMO COMPETITOR PRICE (PCS)": r['p_pcs'],
                        "NORMAL COMPETITOR PRICE (CTN)": r['n_ctn'],
                        "PROMO COMPETITOR PRICE (CTN)": r['p_ctn']
                    }
                    for col_name, val in map_cols.items():
                        if col_name in headers:
                            ws.cell(row=row_num, column=headers.index(col_name)+1).value = val if val > 0 else None
                
                wb.save(FILE_PATH)
                st.success("✅ Excel Updated!")
            
            st.download_button("📂 Download Photos (ZIP)", zip_buffer.getvalue(), f"Photos_{date_inp}.zip")
    else:
        st.error("Database master_harga.xlsx tidak ditemukan di folder database/")
